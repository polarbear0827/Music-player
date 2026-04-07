import os
import logging
import random
import re
import sqlite3
import asyncio
from typing import Optional
import openpyxl
import discord
from discord.ext import commands
from dotenv import load_dotenv

load_dotenv()
log = logging.getLogger(__name__)

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'spongebob.xlsx')

# OpenRouter config
OPENROUTER_BASE_URL = "https://openrouter.ai/api/v1"
OPENROUTER_MODEL = "minimax/minimax-m2.5:free"
GLOBAL_OR_API_KEY = os.getenv("OPENROUTER_API_KEY", "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")


def load_sticker_data() -> list[dict]:
    """
    Load sticker data from spongebob.xlsx.
    - 網址 sheet: C col (caption), H col (url), row 2+
    - K網址 sheet: D col (caption), I col (url), row 2+
    Returns list of {'caption': str, 'url': str, 'series': str}
    """
    stickers = []
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

        # --- 網址 sheet: C=col3 (name/caption), H=col8 (i.imgur url) ---
        if '網址' in wb.sheetnames:
            ws = wb['網址']
            for row in ws.iter_rows(min_row=2, values_only=True):
                caption = row[2] if len(row) > 2 else None   # col C (index 2)
                url = row[7] if len(row) > 7 else None       # col H (index 7)
                if caption and url and isinstance(url, str) and url.startswith('http'):
                    stickers.append({
                        'caption': str(caption).strip(),
                        'url': str(url).strip(),
                        'series': 'SS'
                    })

        # --- K網址 sheet: D=col4 (caption), I=col9 (i.imgur url) ---
        if 'K網址' in wb.sheetnames:
            ws = wb['K網址']
            for row in ws.iter_rows(min_row=2, values_only=True):
                caption = row[3] if len(row) > 3 else None   # col D (index 3)
                url = row[8] if len(row) > 8 else None       # col I (index 8)
                if caption and url and isinstance(url, str) and url.startswith('http'):
                    stickers.append({
                        'caption': str(caption).strip(),
                        'url': str(url).strip(),
                        'series': 'SK'
                    })

        wb.close()
        log.info(f"Sticker DB loaded: {len(stickers)} stickers total.")
    except Exception as e:
        log.error(f"Failed to load sticker data from Excel: {e}")

    return stickers


def search_stickers(stickers: list[dict], keyword: str, limit: int = 5) -> list[dict]:
    """Fuzzy substring search across all captions."""
    kw = keyword.strip().lower()
    # First pass: exact substring
    results = [s for s in stickers if kw in s['caption'].lower()]
    if not results:
        # Second pass: every char of keyword must appear in order (very loose)
        pattern = '.*'.join(re.escape(c) for c in kw)
        results = [s for s in stickers if re.search(pattern, s['caption'].lower())]
    return results[:limit]


class StickerDB:
    """Handles per-user OpenRouter API key storage in SQLite."""

    def __init__(self, db_name: str = 'sticker_keys.db'):
        os.makedirs('data', exist_ok=True)
        import shutil
        if os.path.exists(db_name) and not os.path.exists(os.path.join('data', db_name)):
            try:
                shutil.move(db_name, os.path.join('data', db_name))
                log.info(f"Migrated {db_name} to data folder")
            except Exception as e:
                log.error(f"Failed to migrate db: {e}")

        self.db_path = os.path.join('data', db_name)
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.conn.execute('''
            CREATE TABLE IF NOT EXISTS user_keys (
                user_id INTEGER PRIMARY KEY,
                api_key TEXT NOT NULL
            )
        ''')
        self.conn.commit()

    def set_key(self, user_id: int, api_key: str):
        self.conn.execute(
            'INSERT OR REPLACE INTO user_keys (user_id, api_key) VALUES (?, ?)',
            (user_id, api_key)
        )
        self.conn.commit()

    def get_key(self, user_id: int) -> Optional[str]:
        row = self.conn.execute(
            'SELECT api_key FROM user_keys WHERE user_id = ?', (user_id,)
        ).fetchone()
        return row[0] if row else None

    def delete_key(self, user_id: int):
        self.conn.execute('DELETE FROM user_keys WHERE user_id = ?', (user_id,))
        self.conn.commit()


class StickerSelectView(discord.ui.View):
    """Shown when multiple stickers match the keyword — user picks one."""

    def __init__(self, matches: list[dict], author: discord.User):
        super().__init__(timeout=30)
        self.matches = matches
        self.author = author
        self.chosen: Optional[dict] = None

        for i, sticker in enumerate(matches):
            btn = discord.ui.Button(
                label=f"{i + 1}. {sticker['caption'][:40]}",
                style=discord.ButtonStyle.secondary,
                custom_id=f"sticker_pick_{i}"
            )
            btn.callback = self._make_callback(i)
            self.add_item(btn)

    def _make_callback(self, index: int):
        async def callback(interaction: discord.Interaction):
            if interaction.user.id != self.author.id:
                return await interaction.response.send_message("這不是你的選擇！", ephemeral=True)
            self.chosen = self.matches[index]
            self.stop()
            await interaction.response.defer()
        return callback

    async def on_timeout(self):
        self.stop()


class Sticker(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.sticker_db = StickerDB()
        # Load stickers at startup (blocking but only once)
        self.stickers: list[dict] = []
        bot.loop.create_task(self._async_load_stickers())

    async def _async_load_stickers(self):
        self.stickers = await asyncio.to_thread(load_sticker_data)
        log.info(f"[Sticker] Async load complete: {len(self.stickers)} stickers")

    def _get_api_key(self, user_id: int) -> str:
        """Returns user-specific key or falls back to global."""
        return self.sticker_db.get_key(user_id) or GLOBAL_OR_API_KEY

    @commands.command(name='sticker', aliases=['貼圖', 'ss'], help='搜尋海綿寶寶貼圖。用法: F!sticker <關鍵字>')
    async def sticker(self, ctx: commands.Context, *, keyword: str):
        if not self.stickers:
            return await ctx.send("⚠️ 貼圖資料庫尚未載入，請稍後再試！")

        matches = search_stickers(self.stickers, keyword)

        if not matches:
            return await ctx.send(f"❌ 找不到包含 **{keyword}** 的貼圖，換個關鍵字試試！")

        if len(matches) == 1:
            # Direct hit
            s = matches[0]
            embed = discord.Embed(
                description=f"🪸 **{s['caption']}**",
                color=0xFFD700
            )
            embed.set_image(url=s['url'])
            embed.set_footer(text=f"搜尋: {keyword} | 系列: {s['series']}")
            return await ctx.send(embed=embed)

        # Multiple matches — show selection buttons
        view = StickerSelectView(matches, ctx.author)
        choice_text = "\n".join(
            f"**{i+1}.** {s['caption'][:60]}" for i, s in enumerate(matches)
        )
        msg = await ctx.send(
            f"🔍 找到 **{len(matches)}** 個相關貼圖，請選擇（30秒）：\n{choice_text}",
            view=view
        )

        await view.wait()

        if view.chosen:
            s = view.chosen
            embed = discord.Embed(
                description=f"🪸 **{s['caption']}**",
                color=0xFFD700
            )
            embed.set_image(url=s['url'])
            embed.set_footer(text=f"搜尋: {keyword} | 系列: {s['series']}")
            try:
                await msg.edit(content=None, embed=embed, view=None)
            except discord.NotFound:
                pass
        else:
            try:
                await msg.edit(content="⏰ 選擇超時，已取消。", view=None)
            except discord.NotFound:
                pass

    @commands.command(name='sticker_random', aliases=['隨機貼圖', 'ssr'], help='隨機送出一張海綿寶寶貼圖')
    async def sticker_random(self, ctx: commands.Context):
        if not self.stickers:
            return await ctx.send("⚠️ 貼圖資料庫尚未載入！")

        s = random.choice(self.stickers)
        embed = discord.Embed(
            description=f"🎲 **{s['caption']}**",
            color=0xFF6B9D
        )
        embed.set_image(url=s['url'])
        embed.set_footer(text=f"系列: {s['series']} | 共 {len(self.stickers)} 張圖")
        await ctx.send(embed=embed)

    @commands.command(name='setkey', help='設定你的 OpenRouter API Key（私訊傳送比較安全）')
    async def setkey(self, ctx: commands.Context, *, api_key: str):
        # Delete the message immediately for security (if in a public channel)
        try:
            await ctx.message.delete()
        except Exception:
            pass

        if not api_key.startswith('sk-or-'):
            return await ctx.send("❌ 無效的 OpenRouter API Key 格式（應以 `sk-or-` 開頭）", ephemeral=True)

        self.sticker_db.set_key(ctx.author.id, api_key)
        await ctx.send(
            f"✅ {ctx.author.mention} 的 OpenRouter API Key 已儲存！\n"
            "⚠️ 建議下次在私訊中使用此指令以保護 API Key。",
            delete_after=10
        )

    @commands.command(name='delkey', help='刪除你的 OpenRouter API Key')
    async def delkey(self, ctx: commands.Context):
        self.sticker_db.delete_key(ctx.author.id)
        await ctx.send("🗑️ 你的 OpenRouter API Key 已刪除，將改用預設金鑰。", delete_after=5)

    @commands.command(name='sticker_count', aliases=['貼圖數量'], help='顯示目前貼圖資料庫統計')
    async def sticker_count(self, ctx: commands.Context):
        ss_count = sum(1 for s in self.stickers if s['series'] == 'SS')
        sk_count = sum(1 for s in self.stickers if s['series'] == 'SK')
        embed = discord.Embed(
            title="📊 海綿寶寶貼圖資料庫統計",
            color=0x00BFFF
        )
        embed.add_field(name="🟡 一般系列 (SS)", value=f"`{ss_count}` 張", inline=True)
        embed.add_field(name="🟠 K系列 (SK)", value=f"`{sk_count}` 張", inline=True)
        embed.add_field(name="📦 總計", value=f"`{len(self.stickers)}` 張", inline=True)
        await ctx.send(embed=embed)


async def setup(bot: commands.Bot):
    await bot.add_cog(Sticker(bot))
